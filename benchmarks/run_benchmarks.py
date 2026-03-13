from __future__ import annotations

import argparse
import importlib
import json
import sys
import time
from dataclasses import asdict, dataclass, field
from pathlib import Path
from statistics import median
from typing import Any, Callable
from xml.etree import ElementTree as ET

import fitz

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from axiomdoc.pipeline import parse_to_artifacts


@dataclass
class DatasetEntry:
    document_id: str
    path: str
    doc_type: str
    pages: int | None = None
    labels: dict[str, Any] = field(default_factory=dict)
    source: str | None = None

    @property
    def resolved_path(self) -> Path:
        return ROOT / self.path if not Path(self.path).is_absolute() else Path(self.path)


@dataclass
class DocumentResult:
    document_id: str
    document: str
    doc_type: str
    pages: int
    runtime_seconds: float
    seconds_per_page: float
    xml_well_formed: bool
    markdown_chars: int
    heading_count: int
    chunk_count: int
    success: bool
    error: str | None = None


def run_axiomdoc(path: Path) -> tuple[str, str, dict]:
    _, xml_output, markdown_output, index_manifest = parse_to_artifacts(path)
    return xml_output, markdown_output, index_manifest


def run_pymupdf_raw(path: Path) -> tuple[str, str, dict]:
    with fitz.open(path) as pdf:
        pages: list[str] = []
        for page in pdf:
            pages.append(page.get_text("text"))
        markdown = "\n\n".join(text.strip() for text in pages if text.strip()) + "\n"
    xml_output = f"<document source_name=\"{path.name}\"><text>{_xml_escape(markdown)}</text></document>"
    return xml_output, markdown, {"chunks": []}


def run_docling(path: Path) -> tuple[str, str, dict]:
    converter_module = importlib.import_module("docling.document_converter")
    converter = converter_module.DocumentConverter()
    result = converter.convert(str(path))
    document = result.document
    markdown = document.export_to_markdown()
    xml_output = f"<document source_name=\"{path.name}\"><text>{_xml_escape(markdown)}</text></document>"
    return xml_output, markdown, {"chunks": []}


def run_pdfplumber(path: Path) -> tuple[str, str, dict]:
    pdfplumber = importlib.import_module("pdfplumber")
    with pdfplumber.open(str(path)) as pdf:
        texts = []
        for page in pdf.pages:
            text = page.extract_text() or ""
            if text.strip():
                texts.append(text.strip())
    markdown = "\n\n".join(texts) + "\n"
    xml_output = f"<document source_name=\"{path.name}\"><text>{_xml_escape(markdown)}</text></document>"
    return xml_output, markdown, {"chunks": []}


def run_raw_text(path: Path) -> tuple[str, str, dict]:
    raw = path.read_text(encoding="utf-8", errors="ignore")
    markdown = raw.strip() + ("\n" if raw else "")
    xml_output = f"<document source_name=\"{path.name}\"><text>{_xml_escape(markdown)}</text></document>"
    return xml_output, markdown, {"chunks": []}


RUNNERS: dict[str, Callable[[Path], tuple[str, str, dict]]] = {
    "axiomdoc": run_axiomdoc,
    "pymupdf_raw": run_pymupdf_raw,
    "docling": run_docling,
    "pdfplumber": run_pdfplumber,
    "raw_text": run_raw_text,
}


def load_manifest(path: Path) -> list[DatasetEntry]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    entries = [DatasetEntry(**entry) for entry in payload["documents"]]
    return entries


def evaluate_runner(name: str, entries: list[DatasetEntry], progress_path: Path | None = None) -> dict[str, Any]:
    runner = RUNNERS[name]
    results: list[DocumentResult] = []

    for index, entry in enumerate(entries, start=1):
        path = entry.resolved_path
        pages = entry.pages or infer_page_units(path, entry.doc_type)

        started = time.perf_counter()
        try:
            xml_output, markdown_output, index_manifest = runner(path)
            runtime = time.perf_counter() - started
            results.append(
                DocumentResult(
                    document_id=entry.document_id,
                    document=path.name,
                    doc_type=entry.doc_type,
                    pages=pages,
                    runtime_seconds=runtime,
                    seconds_per_page=runtime / max(pages, 1),
                    xml_well_formed=_is_well_formed_xml(xml_output),
                    markdown_chars=len(markdown_output),
                    heading_count=_count_markdown_headings(markdown_output),
                    chunk_count=len(index_manifest.get("chunks", [])),
                    success=True,
                )
            )
        except Exception as exc:
            runtime = time.perf_counter() - started
            results.append(
                DocumentResult(
                    document_id=entry.document_id,
                    document=path.name,
                    doc_type=entry.doc_type,
                    pages=pages,
                    runtime_seconds=runtime,
                    seconds_per_page=runtime / max(pages, 1),
                    xml_well_formed=False,
                    markdown_chars=0,
                    heading_count=0,
                    chunk_count=0,
                    success=False,
                    error=f"{type(exc).__name__}: {exc}",
                )
            )

        if progress_path is not None and (index % 5 == 0 or index == len(entries)):
            progress_payload = {
                "library": name,
                "completed_documents": index,
                "total_documents": len(entries),
                "documents": [asdict(result) for result in results],
                "summary": summarize_results(results),
                "by_doc_type": summarize_by_doc_type(results),
            }
            progress_path.parent.mkdir(parents=True, exist_ok=True)
            progress_path.write_text(json.dumps(progress_payload, indent=2), encoding="utf-8")
            print(f"[{name}] {index}/{len(entries)}", flush=True)

    final_payload = {
        "library": name,
        "documents": [asdict(result) for result in results],
        "summary": summarize_results(results),
        "by_doc_type": summarize_by_doc_type(results),
    }
    if progress_path is not None:
        progress_path.parent.mkdir(parents=True, exist_ok=True)
        progress_path.write_text(json.dumps(final_payload, indent=2), encoding="utf-8")
    return final_payload


def summarize_results(results: list[DocumentResult]) -> dict[str, Any]:
    successful = [result for result in results if result.success]
    return {
        "document_count": len(results),
        "success_rate": sum(1 for result in results if result.success) / len(results) if results else 0.0,
        "median_seconds_per_page": median(result.seconds_per_page for result in successful) if successful else None,
        "xml_well_formed_rate": (
            sum(1 for result in successful if result.xml_well_formed) / len(successful) if successful else 0.0
        ),
        "median_markdown_chars": median(result.markdown_chars for result in successful) if successful else None,
        "median_heading_count": median(result.heading_count for result in successful) if successful else None,
        "median_chunk_count": median(result.chunk_count for result in successful) if successful else None,
    }


def summarize_by_doc_type(results: list[DocumentResult]) -> dict[str, Any]:
    payload: dict[str, list[DocumentResult]] = {}
    for result in results:
        payload.setdefault(result.doc_type, []).append(result)
    return {doc_type: summarize_results(items) for doc_type, items in sorted(payload.items())}


def infer_page_units(path: Path, doc_type: str) -> int:
    doc_type = doc_type.lower()
    if doc_type == "pdf":
        with fitz.open(path) as pdf:
            return pdf.page_count
    if doc_type in {"docx", "xml", "html", "txt"}:
        return 1
    if doc_type in {"xlsx", "xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return 1
        workbook = load_workbook(filename=str(path), read_only=True)
        count = len(workbook.sheetnames)
        workbook.close()
        return max(count, 1)
    return 1


def _is_well_formed_xml(xml_output: str) -> bool:
    try:
        ET.fromstring(xml_output)
        return True
    except ET.ParseError:
        return False


def _count_markdown_headings(markdown_output: str) -> int:
    return sum(1 for line in markdown_output.splitlines() if line.lstrip().startswith("#"))


def _xml_escape(text: str) -> str:
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&apos;")
    )


def main() -> int:
    parser = argparse.ArgumentParser(description="Run benchmark comparisons for AXIOMDoc against baseline libraries.")
    parser.add_argument("--manifest", type=Path, required=True, help="JSON dataset manifest")
    parser.add_argument("--libraries", nargs="+", default=["axiomdoc", "pymupdf_raw", "pdfplumber", "docling"])
    parser.add_argument("--output", type=Path, default=Path("benchmarks/results/latest.json"))
    parser.add_argument("--progress-dir", type=Path, default=Path("benchmarks/results/progress"))
    args = parser.parse_args()

    entries = load_manifest(args.manifest)
    if not entries:
        raise SystemExit(f"No documents found in manifest {args.manifest}")

    payload = {
        "manifest": str(args.manifest),
        "document_count": len(entries),
        "doc_type_counts": count_doc_types(entries),
        "libraries": [],
    }
    for library in args.libraries:
        progress_path = args.progress_dir / f"{args.output.stem}-{library}.json" if args.progress_dir else None
        payload["libraries"].append(evaluate_runner(library, entries, progress_path=progress_path))

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(args.output)
    return 0


def count_doc_types(entries: list[DatasetEntry]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for entry in entries:
        counts[entry.doc_type] = counts.get(entry.doc_type, 0) + 1
    return dict(sorted(counts.items()))


if __name__ == "__main__":
    raise SystemExit(main())
