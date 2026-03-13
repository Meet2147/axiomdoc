from __future__ import annotations

from pathlib import Path

from axiomdoc.models import Block, CanonicalDocument
from axiomdoc.parsers.base import ParserBackend


class XlsxParser(ParserBackend):
    name = "openpyxl"
    supported_suffixes = (".xlsx", ".xlsm")

    def parse(self, path: Path) -> CanonicalDocument:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "XLSX parsing requires the 'openpyxl' dependency. Install 'axiomdoc[xlsx]' or 'axiomdoc[full]'."
            ) from exc

        workbook = load_workbook(filename=str(path), data_only=True, read_only=True)
        document = CanonicalDocument.empty(path, source_format="xlsx", parser_name=self.name)
        document.metadata["sheet_count"] = len(workbook.sheetnames)

        block_counter = 1
        for sheet in workbook.worksheets:
            document.blocks.append(
                Block(
                    block_id=f"s{block_counter}",
                    kind="heading",
                    text=sheet.title,
                    level=1,
                    metadata={"sheet": sheet.title},
                )
            )
            block_counter += 1

            non_empty_rows = 0
            table_rows: list[list[str]] = []
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                if not values:
                    continue

                if len(values) > 1:
                    table_rows.append(values)
                else:
                    document.blocks.append(
                        Block(
                            block_id=f"s{block_counter}",
                            kind="paragraph",
                            text=values[0],
                            metadata={"sheet": sheet.title, "row_index": row_index},
                        )
                    )
                    block_counter += 1
                non_empty_rows += 1

            if table_rows:
                document.blocks.append(
                    Block(
                        block_id=f"s{block_counter}",
                        kind="table",
                        text="\n".join(" | ".join(row) for row in table_rows),
                        metadata={
                            "sheet": sheet.title,
                            "table_rows": table_rows,
                            "column_count": max(len(row) for row in table_rows),
                        },
                    )
                )
                block_counter += 1

            if non_empty_rows == 0:
                document.blocks.append(
                    Block(
                        block_id=f"s{block_counter}",
                        kind="paragraph",
                        text="(empty sheet)",
                        metadata={"sheet": sheet.title},
                    )
                )
                block_counter += 1

        workbook.close()
        return document
